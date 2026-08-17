# -*- coding: utf-8 -*-
"""企业预警通 通用抓取 CLI — 抓取网站上任意可获取的信息

三种抓取对象(信息维度不限, 由入口发现器动态支持, 不硬编码):
  1. 企业详情页的任意入口: --entry "债券融资" / "股东" / "风险" ... (模糊匹配, 实时枚举)
  2. 企业全部入口:         --all (逐个抓取, 注意配额)
  3. 任意站内 URL:         --url https://www.qyyjt.cn/... (搜索页/榜单页/公告页等)

用法:
  python qyyjt_fetch.py "企业名" --list                 # 列出该企业所有可用入口
  python qyyjt_fetch.py "企业名" --entry "债券融资"      # 抓指定入口(中文/英文别名/部分匹配)
  python qyyjt_fetch.py "企业名" --entry 股东 --all-rows # 抓入口并保留完整行数据
  python qyyjt_fetch.py "企业名" --all                  # 抓全部入口(耗配额, 谨慎)
  python qyyjt_fetch.py --url "https://www.qyyjt.cn/s?tab=securities&k=重庆"  # 任意 URL
  python qyyjt_fetch.py "企业名" --entry "债券融资" --out result.xlsx
  python qyyjt_fetch.py "企业名" --entry "债券融资" --out result.json --full-api
  python qyyjt_fetch.py "企业名" --entry "对外投资企业" --max-pages 5  # 翻页合并

输出: 默认 stdout 打印摘要; --out .json 存结构化结果; --out .xlsx 存多 Sheet Excel。
退出码: 0=成功 2=配额停止 3=未登录 4=未找到/入口不存在 5=权限不足(需正式会员)
"""
import argparse
import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from qyyjt_common import (  # noqa: E402
    ApiCapture, EXIT_LOGIN, EXIT_NOTFOUND, EXIT_OK, EXIT_PERM, EXIT_QUOTA,
    PermissionDenied, QuotaExceeded, NotLoggedIn, account_id, check_permission,
    click_entry, click_path, close_browser, collect_entries, collect_tree,
    expand_tree_for_keyword, expand_tree_nodes, extract_blocks, extract_tables,
    flatten_tree, goto_overview, open_browser, page_text, pick_best,
    search_company, stat_lines, tree_fingerprint,
)
from branch_nav import (BranchNavigator, MatrixParser, ParamRewriter,
                        locate_path_fuzzy)  # noqa: E402

DATA_DIR = Path(__file__).parent.parent / 'data'
ALIASES = {
    # 英文/拼音别名 -> 中文子串(匹配入口文本)
    'overview': '概览', 'basic_info': '基本信息', 'basic': '基本信息',
    'shareholder': '股东', 'stock': '股东', 'guquan': '股东',
    'financial': '财务', 'finance': '财务', 'caiwu': '财务',
    'risk': '风险', 'lawsuit': '诉讼', 'judicial': '司法',
    'admin': '处罚', 'abnormal': '异常', 'credit': '信用',
    'rating': '评级', 'bond': '债券', 'zhaiq': '债券',
    'financing': '融资', 'rongzi': '融资', 'graph': '图谱',
    'equity': '股权', 'change': '变更', 'operation': '经营',
    'bidding': '招投标', 'ip': '知识产权', 'patent': '专利',
    'trademark': '商标', 'news': '新闻', 'notice': '公告',
}


def log(msg):
    print(msg, flush=True)


def print_tree(nodes, indent=0):
    """树形打印入口菜单"""
    for n in nodes:
        if n.get('expandable'):
            icon = '▾' if n.get('sw') == 'open' else '▸'
        else:
            icon = '·'
        log('  ' * indent + f'{icon} {n["text"]}')
        if n.get('children'):
            print_tree(n['children'], indent + 1)


def match_entry(entries, query, fuzzy=True, prefer_leaf=True):
    """入口匹配: 精确 -> 子串(入口含查询词优先) -> 别名 -> 模糊(可选)。

    prefer_leaf=True 时在同等匹配中优先返回非可展开入口(数据入口), 避免
    命中 tree 父节点(如'司法诉讼'树)这种"只展开不加载数据"的节点。
    返回 entry 或 None。
    """
    q = (query or '').strip()
    if not q:
        return None
    candidates = []
    for e in entries:
        if e['text'] == q:
            candidates.append(e)
            break
    if not candidates:
        for e in entries:
            if q in e['text']:
                candidates.append(e)
    if not candidates:
        for e in entries:
            if e['text'] in q:
                candidates.append(e)
    if not candidates:
        alias = ALIASES.get(q.lower())
        if alias:
            for e in entries:
                if alias in e['text'] or e['text'] in alias:
                    candidates.append(e)
    if not candidates and fuzzy:
        import difflib
        best = difflib.get_close_matches(q, [e['text'] for e in entries], n=1, cutoff=0.66)
        if best:
            for e in entries:
                if e['text'] == best[0]:
                    candidates.append(e)
    if not candidates:
        return None
    if prefer_leaf:
        for e in candidates:
            if not e.get('expandable'):
                return e
    return candidates[0]


def summarize(rec):
    """结果记录 -> 打印摘要"""
    if rec.get('permission_denied'):
        log(f"  [{rec.get('entry','?')}] !! 权限不足: {rec.get('permission_msg','')[:120]}")
        return
    log(f"  [{rec.get('entry','?')}]")
    apis = rec.get('api', [])
    if apis:
        for a in apis:
            keys = ','.join(a['keys'][:6])
            log(f"    API {a['code']}  (命中{a['count']}次, 数据{a['rowCount']}行, 键: {keys})")
    for t in rec.get('tables', []):
        pages = f" (已翻页合并)" if t.get('paginated') else ""
        log(f"    表格: {len(t['headers'])}列 {t['rowCount']}行{pages}  表头: {','.join(t['headers'][:8])}")
    blocks = rec.get('blocks', [])
    if blocks:
        kvs = [b for b in blocks if b['type'] == 'kv']
        lists = [b for b in blocks if b['type'] == 'list']
        cards = [b for b in blocks if b['type'] == 'card']
        log(f"    内容块: 键值对{kvs and len(kvs) or 0} 列表{lists and len(lists) or 0} 卡片{cards and len(cards) or 0}"
            + (f"  示例: {lists[0]['value'][:60]}" if lists else ""))
    st = rec.get('stats', [])
    if st:
        log(f"    统计: {' | '.join(st[:6])}")


def to_excel(out_path, results):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '概览'
    ws.append(['公司', '入口', 'API端点', '表格数', '总行数', '矩阵数', '抓取时间'])
    for r in results:
        n_tables = len(r.get('tables', []))
        n_rows = sum(t['rowCount'] for t in r.get('tables', []))
        n_matrix = len(r.get('matrix', []))
        api = '; '.join(a['code'] for a in r.get('api', [])[:3])
        ws.append([r.get('company', ''), r.get('entry', ''), api, n_tables, n_rows,
                   n_matrix, r.get('time', '')])
    for i, r in enumerate(results, start=2):
        # 报表矩阵优先(结构化更好)
        for mi, m in enumerate(r.get('matrix', [])):
            title = f"{r.get('entry', '报表')}_{mi + 1}"[:31]
            MatrixParser.to_excel_sheet(wb, title, m,
                                        meta={'title': r.get('entry', ''),
                                              'unit': '见页面'})
        for t in r.get('tables', []):
            name = f"{r.get('entry', 'entry')}_表{t.get('table', i)}"[:31] or 'sheet'
            s = wb.create_sheet(title=name)
            s.append(t['headers'])
            for row in t['rows']:
                s.append(row)
            for c in s[1]:
                c.font = Font(bold=True)
    wb.save(str(out_path))
    log(f'已写入 Excel: {out_path}  ({len(results)} 个入口, {len(wb.sheetnames)} 个 Sheet)')


async def fetch_entry(pg, entry, company_name, keep_full=False, wait_ms=5000,
                      max_pages=0, clicked=False):
    """抓取入口数据: API 摘要 + 表格(可翻页) + 内容块 + 统计 + 正文快照。

    clicked=True 时跳过点击(调用方已用 click_path 导航到位)。
    权限处理: 点击后若页面出现"权限不足/成为正式用户"等提示, 返回
    permission_denied=True 的记录(不把残留表格当作正常数据)。
    """
    cap = ApiCapture(pg, keep_full=keep_full)
    if not clicked:
        ok = await click_entry(pg, entry, wait_ms=wait_ms, auto_wait=True)
        if not ok:
            return None
    await cap.drain()
    # 权限识别: 必须在提取数据之前, 避免残留表格被当成结果
    try:
        await check_permission(pg, f"入口[{entry.get('text', '')}]")
    except PermissionDenied as ex:
        return {
            'company': company_name, 'entry': entry.get('text', ''),
            'kind': entry.get('kind', ''),
            'time': datetime.now().isoformat(timespec='seconds'),
            'permission_denied': True, 'permission_msg': str(ex),
            'api': cap.summary(), 'tables': [], 'blocks': [], 'stats': [],
            'text_snapshot': await page_text(pg, 3000),
        }
    tables = await extract_tables(pg, paginate=max_pages > 0, max_pages=max_pages or 5)
    blocks = await extract_blocks(pg)
    stats = await stat_lines(pg, 20)
    try:
        matrix = await MatrixParser.extract(pg)
    except Exception:
        matrix = []
    rec = {
        'company': company_name,
        'entry': entry.get('text', ''),
        'kind': entry.get('kind', ''),
        'path': entry.get('path'),
        'time': datetime.now().isoformat(timespec='seconds'),
        'api': cap.summary(),
        'tables': tables,
        'matrix': matrix,
        'blocks': blocks,
        'stats': stats,
        'text_snapshot': await page_text(pg, 4000),
    }
    if keep_full:
        rec['api_full'] = cap.full()
    return rec


async def cascade_find_entry(pg, query, code, max_parents=4):
    """级联查找入口: 当前详情页匹配失败时, 依次点击含关键词片段的父入口
    (menu/tree), 在新页面重新枚举再匹配(锚点/tab 型入口如 '债券融资' 藏在
    '融资速览' 页面里)。返回 (最终入口列表, 匹配入口) 或 (None, None)。

    下沉处理: 若匹配到的入口是可展开的树父节点(如'司法诉讼'/'企业融资'),
    点击展开后重新枚举并继续用同关键词匹配——真正加载数据的是展开后出现的
    menu 项(如'司法案件 14')。同一入口只展开一次, 防止死循环。
    """
    entries = await collect_entries(pg)
    e = match_entry(entries, query, fuzzy=False)
    # ── 下沉: 可展开树父节点 -> 展开 -> 继续匹配 ──
    expanded = set()
    for _ in range(3):
        if e is None or not e.get('expandable'):
            break
        key = (e['kind'], e['text'])
        if key in expanded:
            log(f'    入口[{e["text"]}] 已展开过且无更深入口, 停止下沉')
            break
        expanded.add(key)
        log(f'    入口[{e["text"]}] 为可展开树节点, 点击展开后继续下沉匹配...')
        try:
            await click_entry(pg, e, wait_ms=2500)
            entries = await collect_entries(pg)
        except QuotaExceeded as ex:
            log(f'    配额: {ex}')
            break
        e = match_entry(entries, query, fuzzy=False)
        if e is not None and not e.get('expandable'):
            log(f'    下沉命中数据入口: [{e["kind"]}] {e["text"]}')
            return entries, e
    if e is not None:
        return entries, e

    # ── 级联: 点击含关键词片段的父入口, 在新页面匹配(锚点/tab 型) ──
    cands = [c for c in [query, query[-2:], query[:2], query[-3:], query[:3]] if c]
    parents, seen = [], set()
    for c in cands:
        for x in entries:
            k = (x['kind'], x['text'])
            if k in seen:
                continue
            if x['kind'] in ('menu', 'tree') and (c in x['text'] or x['text'] in c):
                seen.add(k)
                parents.append(x)
    if not parents:
        parents = [x for x in entries if x['kind'] == 'menu'][:max_parents]
    parents = parents[:max_parents]
    log(f'    级联候选父入口: {[p["text"] for p in parents]}')

    for parent in parents:
        try:
            if parent['kind'] == 'tree':
                entries2 = await expand_tree_for_keyword(pg, parent['text'])
            else:
                await click_entry(pg, parent, wait_ms=4000)
                entries2 = await collect_entries(pg)
        except QuotaExceeded as ex:
            log(f'    配额/异常: {ex}')
            break
        e2 = match_entry(entries2, query, fuzzy=False)
        if e2 is not None:
            return entries2, e2
        # 回到详情页再试下一个父入口
        await goto_overview(pg, code, wait_menu=False)
    return None, None


async def do_fetch(args):
    p, b, pg = await open_browser(headless=not args.headed, profile=args.profile)

    # ── 任意 URL 模式 ──
    if args.url:
        log(f'=== 抓取 URL: {args.url} ===')
        cap = ApiCapture(pg, keep_full=args.full_api)
        await pg.goto(args.url, wait_until='domcontentloaded', timeout=25000)
        await pg.wait_for_timeout(4000)
        await cap.drain()
        try:
            await check_permission(pg, 'URL')
        except PermissionDenied as ex:
            rec = {'company': '', 'entry': args.url, 'kind': 'url',
                   'time': datetime.now().isoformat(timespec='seconds'),
                   'permission_denied': True, 'permission_msg': str(ex),
                   'api': cap.summary(), 'tables': [], 'blocks': [], 'stats': [],
                   'text_snapshot': await page_text(pg, 3000)}
            summarize(rec)
            await close_browser(p, b)
            return finish(args, [rec]) if args.out else EXIT_PERM
        tables = await extract_tables(pg, paginate=args.max_pages > 0,
                                      max_pages=args.max_pages or 5)
        blocks = await extract_blocks(pg)
        stats = await stat_lines(pg, 30)
        rec = {'company': '', 'entry': args.url, 'kind': 'url',
               'time': datetime.now().isoformat(timespec='seconds'),
               'api': cap.summary(), 'tables': tables, 'blocks': blocks,
               'stats': stats, 'text_snapshot': await page_text(pg, 6000)}
        if args.full_api:
            rec['api_full'] = cap.full()
        results = [rec]
        summarize(rec)
        await close_browser(p, b)
        return finish(args, results)

    # ── 企业模式 ──
    results = await search_company(pg, args.company)
    if not results:
        log('!! 搜索无结果')
        await close_browser(p, b)
        return EXIT_NOTFOUND
    best = pick_best(args.company, results)
    if best is None:
        log(f'!! 无相似匹配, 候选: {[r["name"] for r in results[:3]]}')
        await close_browser(p, b)
        return EXIT_NOTFOUND
    code, matched = best['code'], best['name']
    log(f'匹配企业: {matched} (code={code})')
    await goto_overview(pg, code)

    # ── --scan: 结构扫描(主干-分支清单, 不点叶子, 省配额) ──
    if args.scan:
        from branch_nav import BranchNavigator as _Nav
        nav = _Nav(pg, log)
        scan_paths, seen = [], set()
        dir_count = [0]

        async def scan_reset():
            await goto_overview(pg, code)
            for _ in range(12):
                n = await pg.evaluate(
                    '() => document.querySelectorAll(".ant-tree-treenode").length')
                if n > 10:
                    return
                await pg.wait_for_timeout(1000)

        async def scan_dir(path):
            if dir_count[0] >= 30:
                return
            dir_count[0] += 1
            e = await nav.navigate(path, wait_ms=2500)
            if e is None:
                log(f'  !! 目录展开失败: {" / ".join(path)} (站点端未响应)')
                scan_paths.append({'path': path, 'type': 'dir', 'expand_failed': True})
                return
            tree = await collect_tree(pg)
            root = locate_path_fuzzy(tree, path)
            if root is None:
                return
            for child in root.get('children', []):
                p = path + [child['text']]
                key = ' / '.join(p)
                if key in seen:
                    continue
                seen.add(key)
                if child.get('expandable'):
                    scan_paths.append({'path': p, 'type': 'dir'})
                    await scan_dir(p)
                else:
                    scan_paths.append({'path': p, 'type': 'leaf'})

        def collect_visible(ns, prefix):
            for n in ns:
                p = prefix + [n['text']]
                key = ' / '.join(p)
                if key in seen:
                    continue
                seen.add(key)
                if n.get('expandable'):
                    scan_paths.append({'path': p, 'type': 'dir'})
                    collect_visible(n.get('children', []), p)
                else:
                    scan_paths.append({'path': p, 'type': 'leaf'})

        log('=== 结构扫描(逐目录展开, 约 10-12 次查询/2-3 分钟) ===')
        collect_visible(await collect_tree(pg), [])
        tree0 = await collect_tree(pg)
        for child in tree0:
            if not child.get('expandable'):
                continue
            await scan_reset()
            await scan_dir([child['text']])

        fp = await tree_fingerprint(pg)
        acct = await account_id(pg)
        scan_data = {
            'schema': 'v4',
            'subjectType': 'company',
            'subjectName': matched,
            'code': code,
            'accountId': acct,
            'treeFingerprint': fp,
            'probedAt': datetime.now().isoformat(timespec='seconds'),
            'paths': scan_paths,
        }
        out = Path(args.out) if args.out else DATA_DIR / f'scan_{matched}.json'
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, 'w', encoding='utf-8') as f:
            json.dump(scan_data, f, ensure_ascii=False, indent=1)
        dirs = [p for p in scan_paths if p['type'] == 'dir']
        leaves = [p for p in scan_paths if p['type'] == 'leaf']
        log(f'扫描完成: {len(dirs)} 目录 + {len(leaves)} 叶子 = {len(scan_paths)} 条路径')
        log('=== 主干-分支清单 ===')
        for p in scan_paths:
            mark = ' [展开失败]' if p.get('expand_failed') else ''
            log(f'  {"D " if p["type"]=="dir" else "  "}{" / ".join(p["path"])}{mark}')
        log(f'已保存: {out}  (账号 {acct}, 树指纹 {fp})')
        log('下一步: python qyyjt_fetch.py --open "企业名" --entry "目录/叶子"')
        await close_browser(p, b)
        return EXIT_OK

    entries = await collect_entries(pg)
    log(f'详情页入口 {len(entries)} 个')

    # ── 列出入口(树形展示) ──
    if args.list:
        tree = await collect_tree(pg)
        log('=== 左侧菜单树 ===')
        print_tree(tree)
        extras = [e for e in entries if e['kind'] in ('anchor', 'tab')]
        if extras:
            log('=== 页面级入口(锚点/Tab, 随当前页面变化) ===')
            for e in extras:
                log(f'  [{e["kind"]}] {e["text"]}')
        log('\n用法示例:')
        log(f'  python qyyjt_fetch.py "{args.company}" --entry 债券融资')
        log(f'  python qyyjt_fetch.py "{args.company}" --entry "财务数据/资产负债表"')
        await close_browser(p, b)
        return EXIT_OK

    # ── 路径导航模式: "财务数据/资产负债表" ──
    if args.entry and '/' in args.entry:
        path = [s.strip() for s in args.entry.split('/') if s.strip()]
        log(f'=== 路径导航: {" / ".join(path)} ===')
        # --open: scan 缓存校验(账号指纹 + 树指纹 + 路径存在性)
        if args.open:
            scan_file = Path(args.map) if args.map else DATA_DIR / f'scan_{matched}.json'
            if scan_file.exists():
                try:
                    scan = json.load(open(scan_file, encoding='utf-8'))
                    acct = await account_id(pg)
                    if scan.get('accountId') and acct and scan['accountId'] != acct:
                        log(f'!! 警告: scan 缓存为账号 {scan["accountId"][-11:]} 采集, '
                            f'当前 {acct[-11:]}——菜单为账号自定义, 可能不一致')
                    fp = await tree_fingerprint(pg)
                    if scan.get('treeFingerprint') and scan['treeFingerprint'] != fp:
                        log('!! 警告: 菜单结构与 scan 缓存不一致(改版/账号差异), 建议重扫 --scan')
                    else:
                        log('√ scan 缓存校验通过(账号+树结构一致)')
                    if path not in [p['path'] for p in scan.get('paths', [])]:
                        log(f'!! scan 清单中无此路径, 可用路径:')
                        for p in scan.get('paths', [])[:20]:
                            log(f'    {"D " if p["type"]=="dir" else "  "}{" / ".join(p["path"])}')
                        log('  (仍尝试实时导航, 若失败说明该账号下确实无此入口)')
                except Exception as ex:
                    log(f'!! scan 文件读取失败: {ex}')
            else:
                log(f'!! 未找到 scan 缓存 {scan_file.name}; 建议先 --scan 扫描(仍尝试实时导航)')
        # 参数改写: --params 优先, 其次 --map 中该分支的 paramTemplate
        params = ParamRewriter.parse_params(args.params) if args.params else {}
        if args.map:
            try:
                map_data = json.load(open(args.map, encoding='utf-8'))
                if map_data.get('subjectName') == matched:
                    for b in map_data.get('branches', []):
                        if b.get('path') == path and b.get('paramTemplate'):
                            params = {**b['paramTemplate'], **params}
                            log(f'    地图命中参数模板: {params}')
            except Exception as ex:
                log(f'!! 地图加载失败: {ex}')
        rw = ParamRewriter(pg, params) if params else None
        nav = BranchNavigator(pg, log)
        e = await nav.navigate(path, wait_ms=args.wait)
        if e is None:
            log(f'!! 路径定位失败: {args.entry}; 可用树入口见 --list')
            await close_browser(p, b)
            return EXIT_NOTFOUND
        if rw and rw.rewritten:
            log(f'    参数改写 {len(rw.rewritten)} 次: {rw.rewritten[-1]["from"][:60]} -> {rw.rewritten[-1]["to"][:60]}')
        log(f'=== 抓取入口 [{e["kind"]}] {path[-1]} (路径: {" / ".join(path)}) ===')
        rec = await fetch_entry(pg, {'kind': 'tree', 'text': path[-1], 'path': path},
                                matched, keep_full=args.full_api,
                                max_pages=args.max_pages, clicked=True)
        if rec is None:
            log('!! 抓取失败')
            await close_browser(p, b)
            return EXIT_ERR
        summarize(rec)
        results = [rec]
        if rec.get('permission_denied'):
            await close_browser(p, b)
            finish(args, results)
            return EXIT_PERM

    # ── 抓指定入口 ──
    elif args.entry:
        # 统一走级联: 直接匹配 -> 可展开树节点下沉 -> 父入口级联
        log('定位入口: 直接匹配/树节点下沉/父入口级联...')
        try:
            entries, e = await cascade_find_entry(pg, args.entry, code)
        except QuotaExceeded as ex:
            log(f'!! 级联查找触发配额: {ex}')
            await close_browser(p, b)
            return EXIT_QUOTA
        if e is None and args.expand:
            log('--expand: 全量展开树菜单...')
            try:
                tree = await expand_tree_nodes(pg)
                entries = flatten_tree(tree)
                e = match_entry(entries, args.entry, fuzzy=False)
            except QuotaExceeded as ex:
                log(f'!! 展开树触发配额: {ex}')
        if e is None:
            e = match_entry(entries, args.entry, fuzzy=True)
        if e is None:
            log(f'!! 未找到入口匹配: {args.entry}; 可用入口: {[x["text"] for x in entries]}')
            await close_browser(p, b)
            return EXIT_NOTFOUND
        log(f'=== 抓取入口 [{e["kind"]}] {e["text"]} ===')
        rec = await fetch_entry(pg, e, matched, keep_full=args.full_api,
                                wait_ms=args.wait, max_pages=args.max_pages)
        if rec is None:
            log('!! 点击入口失败')
            await close_browser(p, b)
            return EXIT_ERR
        summarize(rec)
        results = [rec]
        # 单入口权限不足: 结果已标记, 退出码 5 明确告知
        if rec.get('permission_denied'):
            await close_browser(p, b)
            finish(args, results)
            return EXIT_PERM

    # ── 抓全部入口 ──
    elif args.all:
        log(f'=== 抓取全部 {len(entries)} 个入口 (注意配额) ===')
        status = EXIT_OK
        denied = 0
        for i, e in enumerate(entries, 1):
            log(f'[{i}/{len(entries)}] 点击 ({e["kind"]}) {e["text"]} ...')
            try:
                rec = await fetch_entry(pg, e, matched, keep_full=args.full_api,
                                        wait_ms=args.wait, max_pages=args.max_pages)
                if rec:
                    summarize(rec)
                    results.append(rec)
                    if rec.get('permission_denied'):
                        denied += 1
            except QuotaExceeded as ex:
                log(f'!!! 配额停止: {ex}')
                status = EXIT_QUOTA
                break
            # 回到详情页
            await goto_overview(pg, code, wait_menu=False)
        if denied and status == EXIT_OK:
            log(f'!! 其中 {denied} 个入口因权限不足未取到数据')
        if status != EXIT_OK:
            await close_browser(p, b)
            if args.out:
                finish(args, results)
            return status

    else:
        log('!! 需要 --entry / --all / --list / --url 之一')
        await close_browser(p, b)
        return EXIT_OK

    await close_browser(p, b)
    return finish(args, results)


def finish(args, results):
    if args.out:
        out = Path(args.out)
        if out.suffix.lower() == '.xlsx':
            to_excel(out, results)
        else:
            out.parent.mkdir(parents=True, exist_ok=True)
            with open(out, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=1)
            log(f'已写入: {out}')
    return EXIT_OK


async def main():
    ap = argparse.ArgumentParser(description='企业预警通 通用抓取 CLI')
    ap.add_argument('company', nargs='?', help='企业名')
    ap.add_argument('--entry', default=None, help='入口: 中文子串/英文别名(如 债券融资/bond/shareholder)')
    ap.add_argument('--list', action='store_true', help='列出该企业所有入口')
    ap.add_argument('--scan', action='store_true',
                    help='结构扫描: 逐目录展开生成主干-分支清单(约10-12次查询), 保存 data/scan_<企业>.json')
    ap.add_argument('--open', action='store_true',
                    help='(配 --entry 路径) scan 缓存校验后路径直达: 校验账号/树指纹/路径存在性')
    ap.add_argument('--all', action='store_true', help='抓取全部入口(耗配额)')
    ap.add_argument('--expand', action='store_true', help='匹配失败时全量展开树菜单(耗配额)')
    ap.add_argument('--url', default=None, help='直接抓取任意站内 URL')
    ap.add_argument('--out', default=None, help='输出文件 (.json/.xlsx)')
    ap.add_argument('--full-api', action='store_true', help='保留完整 API JSON')
    ap.add_argument('--wait', type=int, default=5000, help='点击后等待毫秒数(默认5000)')
    ap.add_argument('--max-pages', type=int, default=0,
                    help='表格翻页上限(默认0=不翻页; 如 5=每表最多翻到5页; 翻页耗查询配额)')
    ap.add_argument('--params', default=None,
                    help='筛选参数(路径模式): 报告期=2025年报&合并=合并期末&单位=万元')
    ap.add_argument('--map', default=None,
                    help='分支地图 site_map.json(v2), 命中分支自动复用其参数模板')
    ap.add_argument('--profile', default=None, help='浏览器 profile 目录')
    ap.add_argument('--headed', action='store_true', help='有头模式(调试)')
    args = ap.parse_args()

    if not args.company and not args.url:
        log('用法: python qyyjt_fetch.py "企业名" --entry 入口  |  --list  |  --all  |  --scan  |  --open --entry 路径  |  --url <URL>')
        return EXIT_OK
    if args.open and not (args.entry and '/' in args.entry):
        log('!! --open 需要配合路径格式的 --entry: --open --entry "财务数据/主要财务指标"')
        return EXIT_OK
    if args.company and not (args.entry or args.list or args.all or args.scan):
        log('!! 企业模式下需要 --entry / --list / --all / --scan 之一')
        return EXIT_OK

    try:
        return await do_fetch(args)
    except NotLoggedIn as ex:
        log(f'!! {ex}')
        return EXIT_LOGIN
    except QuotaExceeded as ex:
        log(f'!!! {ex}')
        return EXIT_QUOTA
    except PermissionDenied as ex:
        log(f'!! 权限不足: {ex}')
        return EXIT_PERM
    except Exception as ex:
        log(f'!! 异常 {type(ex).__name__}: {ex}')
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(asyncio.run(main()))
