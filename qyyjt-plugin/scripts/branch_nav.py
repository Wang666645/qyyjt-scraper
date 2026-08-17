# -*- coding: utf-8 -*-
"""branch_nav.py — 分支导航器 / 参数改写器 / 报表矩阵解析器 (方案 P1+P2)

P1 BranchNavigator:
  按路径(分支)逐级导航: 展开目录 -> 点击叶子; 每步等待网络静默;
  内容签名比对检测点击是否生效, 失败自动重试; check_quota/check_permission 全程生效。

P1 ParamRewriter:
  拦截指定 API 改写 query 参数后放行(页面上下文请求自动带签名头);
  --params "报告期=2025年报&合并=合并期末" 的中文参数名 -> API 参数映射表。

P2 MatrixParser:
  报表多期矩阵(head/多期列/层级缩进/加粗) -> [{name, indent, bold, values}];
  科学计数法/千分位清洗; Excel 模板输出(标题/单位/缩进/加粗/千分位)。
"""
import json
import re
from urllib.parse import urlparse, parse_qsl, urlencode

from qyyjt_common import (collect_tree, locate_path, click_entry, check_quota,
                          check_permission, PermissionDenied, QuotaExceeded)


# ═══════════════════════════════════════════════════════════
# P1: BranchNavigator
# ═══════════════════════════════════════════════════════════
class BranchNavigator:
    """分支导航执行器: 依序执行路径步骤, 每步失败检测+重试。"""

    def __init__(self, page, log=print):
        self.page = page
        self.log = log

    async def wait_stable(self, timeout=8.0, quiet_ms=600):
        """等待页面稳定: 网络空闲(有限超时) + 静默缓冲"""
        try:
            await self.page.wait_for_load_state('networkidle', timeout=timeout * 1000)
        except Exception:
            pass
        await self.page.wait_for_timeout(quiet_ms)

    async def _signature(self):
        """页面内容签名: 表格数|树节点数|正文长度|正文头部, 用于检测点击是否生效"""
        return await self.page.evaluate("""(function() {
            var tables = document.querySelectorAll('table').length;
            var treeNodes = document.querySelectorAll('.ant-tree-treenode').length;
            var text = (document.body.innerText || '').slice(0, 400);
            return tables + '|' + treeNodes + '|' + text.length + '|' + text.slice(0, 60).replace(/\\s+/g, '');
        })()""")

    async def navigate(self, path, wait_ms=4000, retries=2):
        """按文本路径导航: 中间级逐级展开, 末级点击进入。

        每步点击后比对内容签名: 无变化视为点击未生效, 自动重试(retries 次)。
        中间级若已展开(sw=open)则跳过点击, 避免收起已展开目录。
        返回末级入口 dict 或 None。
        """
        for i in range(len(path) - 1):
            tree = await collect_tree(self.page)
            e = locate_path(tree, path[:i + 1])
            if e is None:
                return None
            if e.get('sw') == 'open':
                continue  # 已展开, 无需点击
            ok = False
            for attempt in range(retries + 1):
                before = await self._signature()
                tree = await collect_tree(self.page)
                e = locate_path(tree, path[:i + 1])
                if e is None:
                    return None
                if not e.get('expandable'):
                    ok = True
                    break
                clicked = await click_entry(self.page,
                                            {'kind': 'tree', 'index': e['index']},
                                            wait_ms=2500)
                await self.wait_stable()
                after = await self._signature()
                if clicked and (after != before):
                    ok = True
                    break
                self.log(f'    展开[{path[i]}] 未生效, 重试 {attempt + 1}/{retries}')
            if not ok:
                self.log(f'!! 展开[{path[i]}] 多次尝试无效果')
                return None
        # 末级: 点击即返回(不校验签名——目标页面可能与当前相同, 签名无变化不代表失败;
        # 加载失败/权限不足由后续 check_quota/check_permission 兜底)
        tree = await collect_tree(self.page)
        e = locate_path(tree, path)
        if e is None:
            return None
        clicked = await click_entry(self.page, {'kind': 'tree', 'index': e['index']},
                                    wait_ms=wait_ms)
        if clicked:
            return e
        return None


# ═══════════════════════════════════════════════════════════
# P1: ParamRewriter (API 参数改写)
# ═══════════════════════════════════════════════════════════
# 中文筛选名 -> API query 参数名(报表类)。运行时可用 --map 中的 paramTemplate 覆盖。
PARAM_ALIASES = {
    '报告期': 'reportDateType', '报告期类型': 'reportDateType',
    '合并': 'mergeRange', '合并范围': 'mergeRange',
    '单位': 'unitCode', '币种': 'unitCode',
    '数据类型': 'dataType', '报表类型': 'dataType',
    '日期': 'date', '类型': 'type',
}


class ParamRewriter:
    """拦截 qyyjt API 请求, 按规则改写 query 参数后放行。

    用法:
        rw = ParamRewriter(page, params={'reportDateType': '1231', 'mergeRange': '1'})
        ... 页面操作 ...
        rw.rewritten  # 实际改写记录
    """

    def __init__(self, page, params=None, url_keywords=None, enabled=True):
        self.page = page
        self.params = params or {}
        self.url_keywords = url_keywords or ['getData.action', 'finchinaAPP']
        self.rewritten = []
        self.enabled = enabled
        if enabled and self.params:
            page.route('**/*', self._route_sync)

    def _route_sync(self, route):
        """sync 包装: playwright 的 route handler 用 task 执行异步逻辑, 避免 coroutine 警告"""
        import asyncio as _aio
        _aio.ensure_future(self._handle(route))

    async def _handle(self, route):
        try:
            url = route.request.url
            if 'qyyjt.cn' not in url:
                await route.continue_()
                return
            if not any(k in url for k in self.url_keywords):
                await route.continue_()
                return
            u = urlparse(url)
            q = dict(parse_qsl(u.query))
            changed = False
            for k, v in self.params.items():
                if k in q:
                    if q[k] != v:
                        q[k] = v
                        changed = True
                else:
                    q[k] = v
                    changed = True
            if not changed:
                await route.continue_()
                return
            new_url = f'{u.scheme}://{u.netloc}{u.path}?' + urlencode(q)
            self.rewritten.append({'from': url[:120], 'to': new_url[:120]})
            await route.continue_(url=new_url)
        except Exception:
            try:
                await route.continue_()
            except Exception:
                pass

    @staticmethod
    def parse_params(spec):
        """'报告期=2025年报&合并=合并期末&单位=万元' -> {'reportDateType':'2025年报',...}

        中文名优先按 PARAM_ALIASES 映射; 无映射时保留原名(API 可能直接用)。
        """
        out = {}
        for seg in (spec or '').split('&'):
            seg = seg.strip()
            if not seg or '=' not in seg:
                continue
            k, v = seg.split('=', 1)
            k, v = k.strip(), v.strip()
            if not k or not v:
                continue
            out[PARAM_ALIASES.get(k, k)] = v
        return out


# ═══════════════════════════════════════════════════════════
# P2: MatrixParser (报表矩阵)
# ═══════════════════════════════════════════════════════════
META_HEADERS = ('报表类型', '单位', '币种', '报告期')


class MatrixParser:
    """报表矩阵解析: 从页面 DOM 提取多期报表(科目名/缩进/加粗/各期数值)。"""

    @staticmethod
    async def extract(page):
        """提取页面中所有报表矩阵(表头含 报告期/报表类型 的 antd 表格)。

        返回 [{headers, rows:[{name, indent, bold, values:[...]}]}]
        - indent: 科目层级(由首列 padding-left 推算, 0=一级科目)
        - bold:   是否加粗(小计/合计行)
        """
        return await page.evaluate("""(function() {
            var out = [];
            document.querySelectorAll('.ant-table-wrapper').forEach(function(w, ti) {
                var headers = [];
                w.querySelectorAll('th').forEach(function(th) {
                    var h = (th.innerText || '').trim().split('\\n')[0];
                    if (h && headers.indexOf(h) < 0) headers.push(h);
                });
                var isReport = headers.some(function(h) {
                    return h.indexOf('报告期') >= 0 || h.indexOf('报表类型') >= 0;
                });
                if (!isReport) return;
                var rows = [];
                w.querySelectorAll('tbody tr').forEach(function(tr) {
                    var tds = tr.querySelectorAll('td');
                    if (!tds.length) return;
                    var first = tds[0];
                    var name = (first.innerText || '').trim().split('\\n')[0].trim();
                    // 清洗私用区图标字符(缩进箭头等 \uE000-\uF8FF)
                    name = name.replace(/[\\uE000-\\uF8FF]/g, '').trim();
                    if (!name) return;
                    var pad = 0;
                    var m = (first.style.paddingLeft || '').match(/(\\d+(?:\\.\\d+)?)px/);
                    if (m) pad = parseFloat(m[1]);
                    var bold = false;
                    first.querySelectorAll('span, div, b, strong').forEach(function(s) {
                        var fw = s.style.fontWeight || '';
                        if (fw === 'bold' || parseInt(fw) >= 600) bold = true;
                        if (s.tagName === 'B' || s.tagName === 'STRONG') bold = true;
                    });
                    var values = [];
                    for (var i = 1; i < tds.length; i++) {
                        values.push((tds[i].innerText || '').trim());
                    }
                    rows.push({name: name, indent: Math.round(pad / 20),
                               bold: bold, values: values});
                });
                out.push({table: ti + 1, headers: headers, rows: rows});
            });
            return out;
        })()""")

    @staticmethod
    def clean_number(v):
        """'1,234.56' -> 1234.56; '--'/'-'/'' -> None; 科学计数法保持字符串"""
        if v is None:
            return None
        s = str(v).strip().replace(',', '')
        if s in ('', '--', '-', 'None', 'nan', 'null'):
            return None
        try:
            return float(s)
        except ValueError:
            return str(v)

    @staticmethod
    def to_excel_rows(matrix, meta=None):
        """矩阵 -> Excel 行: [['科目', *期次列], ...] 含缩进(空格前缀)与加粗标记。

        meta 可选: {'title': 报表名, 'unit': '万元', 'periods': [...]}
        返回 (rows, bold_rows) — bold_rows 为加粗行索引集合。
        """
        headers = matrix.get('headers', [])
        rows = matrix.get('rows', [])
        # 表头: 第一列=科目, 其余=期次(去掉尾部'报表类型'等元信息列)
        head = ['科目'] + [h for h in headers[1:] if h not in META_HEADERS]
        if meta:
            title = meta.get('title', '')
            unit = meta.get('unit', '')
            prefix = []
            if title:
                prefix.append([title] + [''] * (len(head) - 1))
            if unit:
                prefix.append([f'单位: {unit}'] + [''] * (len(head) - 1))
        else:
            prefix = []
        out = list(prefix) + [head]
        bold_rows = {len(prefix)}
        for r in rows:
            name_raw = r['name']
            # 过滤元信息行(截止日期/报表类型等, 表头已含报告期名)
            if any(k in name_raw for k in ('截止日期', '报表类型', '单位', '币种')):
                continue
            values = r.get('values', [])
            # 对齐: values 可能比 head 短(去掉元信息列后)
            vals = values[:max(0, len(head) - 1)]
            name = ('  ' * r.get('indent', 0)) + name_raw
            out.append([name] + vals + [''] * (len(head) - 1 - len(vals)))
            if r.get('bold'):
                bold_rows.add(len(out) - 1)
        return out, bold_rows

    @staticmethod
    def to_excel_sheet(wb, sheet_name, matrix, meta=None):
        """矩阵写入 Excel Sheet(缩进/加粗/千分位数字格式)"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        rows, bold_rows = MatrixParser.to_excel_rows(matrix, meta)
        ws = wb.create_sheet(title=sheet_name[:31])
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                cell = ws.cell(row=ri + 1, column=ci + 1, value=val)
                if ri in bold_rows:
                    cell.font = Font(bold=True)
        # 数字列千分位格式
        for ci in range(2, len(rows[0]) + 1):
            for ri in range(len(rows)):
                cell = ws.cell(row=ri + 1, column=ci)
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'
        ws.column_dimensions['A'].width = 30
        return ws
